import asyncio
from crawl4ai import *
from docx import Document
from openpyxl import load_workbook

async def crawl_and_save(url, index):
    try:
        async with AsyncWebCrawler() as crawler:
            result = await crawler.arun(url=url)
            text_content = result.markdown

            # Tạo file docx riêng cho mỗi URL
            doc = Document()
            doc.add_paragraph(text_content)
            doc.save(f"output_{index}.docx")

            return "OK"
    except Exception as e:
        print(f"Lỗi khi crawl {url}: {e}")
        return "Fail"

async def main():
    # Đọc file template.xlsx
    wb = load_workbook("template.xlsx")
    ws = wb.active

    # Giả sử dòng 1 là header, duyệt từ dòng 2 trở đi
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        url_cell = row[0]   # Cột URL
        status_cell = row[1]  # Cột Status

        url = url_cell.value
        if url is None or url.strip() == "":
            continue  # bỏ qua dòng trống

        status = await crawl_and_save(url, row_index-1)  # index-1 để file tên đẹp hơn
        status_cell.value = status  # cập nhật status vào cột 2

    # Lưu lại file
    wb.save("template.xlsx")
    print("#-> Hoàn thành. Kết quả lưu trong template.xlsx và các file output_x.docx")

if __name__ == "__main__":
    asyncio.run(main())
    try:
        asyncio.run(main())
    except Exception as e:
        print("Có lỗi xảy ra:", e)
    input("Nhấn Enter để thoát...")
